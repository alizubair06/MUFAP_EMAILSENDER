"""
MUFAP AL Habib Daily Rates Pipeline
- Dynamic Category & Sector tracking per fund.
- Mobile-friendly Excel output (Opens properly on Android/iOS/WPS Office).
"""

import os
import re
import sys
import json
import logging
import smtplib
from datetime import datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from typing import Dict, List, Tuple

import pytz
import cloudscraper
from bs4 import BeautifulSoup
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load environment variables
def load_env_file():
    for env_path in [".env", "config.env"]:
        if os.path.exists(env_path):
            with open(env_path, "r", encoding="utf-8") as f:
                for line in f:
                    line = line.strip()
                    if line and not line.startswith("#") and "=" in line:
                        k, v = line.split("=", 1)
                        os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

load_env_file()

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)

MUFAP_URL = "https://mufap.com.pk/Industry/IndustryStatDaily?tab=1"

# Columns exactly as they appear in the live MUFAP table (source order).
# We still need this to correctly index each row's cells, even though we
# no longer output "Sector" or "Benchmark" in the final report.
SOURCE_COLUMNS = [
    "Sector", "Category", "Fund Name", "Rating", "Benchmark",
    "Validity Date", "NAV", "YTD", "MTD", "1 Day",
    "15 Days", "30 Days", "90 Days", "180 Days", "270 Days",
    "365 Days", "2 Years", "3 Years"
]

# Columns we actually want in the Excel/email output -- Sector and
# Benchmark removed, everything else unchanged.
EXACT_HEADERS = [
    "Category", "Fund Name", "Rating",
    "Validity Date", "NAV", "YTD", "MTD", "1 Day",
    "15 Days", "30 Days", "90 Days", "180 Days", "270 Days",
    "365 Days", "2 Years", "3 Years"
]

# ----------------------------------------------------------------------
# 1. Cloudscraper Page Fetch
# ----------------------------------------------------------------------
from seleniumbase import SB

def fetch_mufap_page(url: str) -> str:
    logging.info("Fetching MUFAP page via SeleniumBase (Undetected Chromedriver)...")
    
    max_retries = 3
    for attempt in range(max_retries):
        try:
            # Running headed (will use xvfb in GitHub Actions for full stealth)
            with SB(uc=True, headless=False) as sb:
                sb.get(url)
                # Wait for Cloudflare verification to complete
                sb.sleep(15)
                
                html = sb.get_page_source()
                if "table" not in html.lower():
                    raise ValueError("HTML downloaded but no table found. Cloudflare might still be blocking.")
                
                return html
        except Exception as e:
            logging.warning(f"Attempt {attempt + 1} failed: {e}")
            if attempt == max_retries - 1:
                raise RuntimeError(f"Failed to fetch MUFAP page via SeleniumBase after {max_retries} attempts: {e}")
            import time
            time.sleep(5)

def extract_report_date(soup: BeautifulSoup) -> str:
    text_content = soup.get_text()
    match = re.search(r"(\d{1,2}[-/ ](?:Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[-/ ]\d{4})", text_content, re.IGNORECASE)
    if match:
        raw_date = match.group(1).strip()
        raw_date = raw_date.replace("/", "-").replace(" ", "-")
        parts = raw_date.split("-")
        if len(parts[0]) == 1:
            parts[0] = "0" + parts[0]
        return "-".join(parts).title()
    return datetime.now().strftime("%d-%b-%Y")

# ----------------------------------------------------------------------
# 2. Table Parser (Reads Sector & Category directly from each row)
# ----------------------------------------------------------------------
# NOTE ON THE FIX:
# The MUFAP "Performance Summary" table does NOT use colspan/banner rows to
# switch Sector/Category. Every single <tr> already has its own Sector and
# Category as plain <td> cells, in this exact order:
#
#   Sector | Category | Fund Name | Rating | Benchmark | Validity Date |
#   NAV | YTD | MTD | 1 Day | 15 Days | 30 Days | 90 Days | 180 Days |
#   270 Days | 365 Days | 2 Years | 3 Years
#
# which lines up 1:1 with EXACT_HEADERS. So there is no need to "remember"
# the last-seen category across rows -- that was the bug causing every
# fund to inherit whatever category happened to be detected first
# ("Money Market"). We just read each row's own cells directly.
def parse_mufap_table(soup: BeautifulSoup, keyword: str = "AL Habib") -> Tuple[List[Dict], List[str]]:
    table = soup.find("table")
    if not table:
        tables = soup.find_all("table")
        if tables:
            table = tables[0]
        else:
            raise ValueError("No table found on MUFAP page.")

    rows_data = []
    n_cols = len(SOURCE_COLUMNS)

    body_rows = table.find_all("tr")

    for tr in body_rows:
        tds = tr.find_all("td")  # only real data cells, skip <th> header row
        if not tds:
            continue

        text_cells = [td.get_text(strip=True) for td in tds]
        full_row_str = " ".join(text_cells).strip()

        if not full_row_str:
            continue

        # Skip any leftover header/banner-style rows (safety net only --
        # normal data rows always have >= n_cols cells so this won't trip them)
        if len(text_cells) < n_cols:
            continue

        # Only keep rows for the fund we care about
        if keyword.lower() not in full_row_str.lower():
            continue

        # Map columns 1:1 against SOURCE_COLUMNS (the raw table order),
        # then keep only the fields we actually want in the output
        # (EXACT_HEADERS) -- this drops Sector and Benchmark cleanly.
        full_row_dict = {
            header: text_cells[i] if i < len(text_cells) else "-"
            for i, header in enumerate(SOURCE_COLUMNS)
        }
        row_dict = {header: full_row_dict.get(header, "-") for header in EXACT_HEADERS}

        rows_data.append(row_dict)

    return rows_data, EXACT_HEADERS

# ----------------------------------------------------------------------
# 3. HTML Email Body (Navy Blue Multi-Column Layout)
# ----------------------------------------------------------------------
def build_email_html(data: List[Dict], headers: List[str], report_date: str) -> str:
    html_table = ["<table border='0' cellpadding='0' cellspacing='0' style='border-collapse: collapse; width: max-content; min-width: 60%; margin: 15px auto; font-family: \"Segoe UI\", Tahoma, Geneva, Verdana, sans-serif; font-size: 13px;'>"]
    
    html_table.append("  <tr style='background-color: #1a365d; color: white;'>")
    for h in headers:
        html_table.append(f"    <th style='padding: 10px 15px; text-align: center; white-space: nowrap; border: 1px solid #94a3b8;'>{str(h).strip()}</th>")
    html_table.append("  </tr>")
    
    for i, row in enumerate(data):
        bg_color = "#f8fafc" if i % 2 == 0 else "#ffffff"
        html_table.append(f"  <tr style='background-color: {bg_color};'>")
        for h in headers:
            val = row.get(h, "-")
            html_table.append(f"    <td style='padding: 10px 15px; text-align: center; border: 1px solid #e2e8f0; color: #334155;'>{str(val).strip()}</td>")
        html_table.append("  </tr>")
        
    html_table.append("</table>")
    html_table_str = "\n".join(html_table)

    html = f"""<!DOCTYPE html>
<html>
<head>
  <meta charset="utf-8">
  <title>AL Habib MUFAP Rates</title>
</head>
<body style="margin: 0; padding: 20px; background-color: #f7fafc; font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;">
  <div style="max-width: 900px; margin: 0 auto; background-color: #ffffff; border: 1px solid #e2e8f0; border-radius: 6px; overflow: hidden; box-shadow: 0 4px 6px rgba(0, 0, 0, 0.05);">
    <div style="background-color: #1a365d; color: #ffffff; padding: 18px; font-size: 20px; font-weight: bold; text-align: center; letter-spacing: 0.5px;">
      🚨 NEW MUFAP AL HABIB RATES PUBLISHED
    </div>
    <div style="padding: 20px;">
      <div style="font-size: 14px; color: #718096; margin-bottom: 5px; font-weight: bold; text-transform: uppercase;">
        Report: MUFAP Daily Rates — AL Habib Funds
      </div>
      <div style="font-size: 14px; color: #718096; margin-bottom: 25px; font-weight: bold; text-transform: uppercase;">
        Rates Date: {report_date}
      </div>
      
      <div style="overflow-x: auto; max-width: 100%; border-radius: 4px; padding-bottom: 10px;">
        {html_table_str}
      </div>
      
      <div style="margin-top: 30px; font-size: 11px; color: #a0aec0; line-height: 1.5; border-top: 1px solid #edf2f7; padding-top: 15px; font-style: italic;">
        Disclaimer: This data has been extracted in real-time from the official MUFAP daily sheet. Please refer to the attached Excel file to verify any rates.<br><br>
        <strong>Source URL:</strong> <a href="{MUFAP_URL}" style="color: #3182ce; text-decoration: underline;">{MUFAP_URL}</a>
      </div>
    </div>
  </div>
</body>
</html>
"""
    return html

# ----------------------------------------------------------------------
# 4. Mobile Compatible Excel Generator
# ----------------------------------------------------------------------
def build_excel_file(data: List[Dict], headers: List[str], filename: str = "AL_Habib_MUFAP_Report.xlsx", report_date: str = ""):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "AL Habib Funds"
    ws.views.sheetView[0].showGridLines = True

    # Standard ARGB Hex for Mobile Readers Compatibility
    navy_fill = PatternFill(start_color="FF003366", end_color="FF003366", fill_type="solid")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFFFF")
    data_font = Font(name="Calibri", size=10, color="FF000000")

    border_thin = Side(style="thin", color="FFD3D3D3")
    box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    # Headers
    for col_num, h_name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=h_name)
        cell.fill = navy_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = box_border

    # Data Rows
    for row_num, row_data in enumerate(data, 2):
        for col_num, h_name in enumerate(headers, 1):
            val = row_data.get(h_name, "")

            # Numeric conversion so Excel on mobile displays values clean without errors
            val_to_write = val
            if isinstance(val, str) and val.replace(".", "", 1).replace("-", "", 1).isdigit():
                try:
                    val_to_write = float(val) if "." in val else int(val)
                except ValueError:
                    val_to_write = val

            cell = ws.cell(row=row_num, column=col_num, value=val_to_write)
            cell.font = data_font
            cell.border = box_border
            align_h = "left" if col_num <= 2 else "center"
            cell.alignment = Alignment(horizontal=align_h, vertical="center")

    # Column Auto Width Adjust
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    wb.save(filename)
    logging.info(f"Mobile-compatible Excel report saved: {filename}")

# ----------------------------------------------------------------------
# 5. Email Dispatch Execution
# ----------------------------------------------------------------------
def send_email(subject: str, html_body: str, attachment_path: str):
    smtp_server = os.getenv("SMTP_SERVER", "smtp.gmail.com")
    smtp_port = int(os.getenv("SMTP_PORT", "587"))
    sender_email = os.getenv("SENDER_EMAIL")
    sender_password = os.getenv("SENDER_PASSWORD")
    recipient_emails_raw = os.getenv("RECIPIENT_EMAILS", "")
    recipient_emails = [e.strip() for e in recipient_emails_raw.split(",") if e.strip()]

    if not sender_email or not sender_password or not recipient_emails:
        raise ValueError("Missing email environment variables.")

    msg = MIMEMultipart()
    msg["From"] = sender_email
    msg["To"] = ", ".join(recipient_emails)
    msg["Subject"] = subject

    msg.attach(MIMEText(html_body, "html"))

    if os.path.exists(attachment_path):
        with open(attachment_path, "rb") as f:
            part = MIMEApplication(f.read(), Name=os.path.basename(attachment_path))
            part["Content-Disposition"] = f'attachment; filename="{os.path.basename(attachment_path)}"'
            msg.attach(part)

    with smtplib.SMTP(smtp_server, smtp_port) as server:
        server.starttls()
        server.login(sender_email, sender_password)
        server.send_message(msg)

    logging.info(f"Email successfully sent to {', '.join(recipient_emails)}")


def main():
    logging.info("Starting MUFAP AL Habib Pipeline...")
    
    # PKT Timezone configuration
    pkt_zone = pytz.timezone("Asia/Karachi")
    now_pkt = datetime.now(pkt_zone)
    today_str = now_pkt.strftime("%d-%b-%Y")
    
    state_file = "state.json"
    force_send = os.getenv("FORCE_SEND", "false").lower() in ("true", "1", "yes")
    
    # State tracking: Avoid double-emailing (unless FORCE_SEND is enabled for testing)
    if os.path.exists(state_file) and not force_send:
        try:
            with open(state_file, "r") as f:
                state = json.load(f)
                if state.get("last_run_date") == today_str:
                    logging.info(f"Already processed MUFAP emails for today ({today_str}). Exiting.")
                    sys.exit(0)
        except Exception as e:
            logging.warning(f"Could not read {state_file}: {e}")
    elif force_send:
        logging.info("FORCE_SEND is enabled: Bypassing already-sent check for testing.")

    html_content = fetch_mufap_page(MUFAP_URL)
    soup = BeautifulSoup(html_content, "html.parser")

    report_date = extract_report_date(soup)
    logging.info(f"MUFAP Live Report Date: {report_date}")
    
    # Strict Date Checking (unless FORCE_SEND is enabled for testing)
    if report_date != today_str and not force_send:
        logging.warning(f"MUFAP report date ({report_date}) does not match today's date ({today_str}). The site hasn't updated yet. Exiting.")
        sys.exit(0)
    elif report_date != today_str and force_send:
        logging.info(f"FORCE_SEND is enabled: Processing latest MUFAP report ({report_date}) even though today is {today_str}.")

    data_rows, headers = parse_mufap_table(soup, keyword="AL Habib")
    logging.info(f"Extracted {len(data_rows)} AL Habib funds.")

    if not data_rows:
        logging.warning("No AL Habib data extracted. Exiting.")
        sys.exit(0)

    clean_date = re.sub(r"[^\w\-]", "_", report_date)
    excel_file = f"AL_Habib_MUFAP_{clean_date}.xlsx"

    build_excel_file(data_rows, headers, filename=excel_file, report_date=report_date)
    email_html = build_email_html(data_rows, headers, report_date)

    subject = f"MUFAP Daily Performance Report — AL Habib Funds ({report_date})"
    send_email(subject, email_html, excel_file)
    logging.info("Pipeline completed & email sent successfully!")
    
    # Update State
    try:
        with open(state_file, "w") as f:
            json.dump({"last_run_date": today_str}, f)
        logging.info("Updated state.json.")
    except Exception as e:
        logging.error(f"Failed to update state.json: {e}")


if __name__ == "__main__":
    main()